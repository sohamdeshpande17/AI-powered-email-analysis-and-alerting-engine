"""
Attachment text extraction — extracts text from PDF, Word, Excel, CSV,
and plain-text attachments.  Also handles hyperlinked documents and ZIP files.

No internal project dependencies (uses only third-party libs + stdlib).
"""

import csv
import io
import re
from typing import List, Dict, Any, Tuple
from urllib.parse import urlparse, unquote

import requests as http_requests


# ── Individual Format Extractors ──────────────────────────────────────

def extract_pdf_text(data: bytes) -> str:
    """Extract text from a PDF file."""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(data))
        pages = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"[Page {i+1}]\n{text.strip()}")
        return "\n\n".join(pages) if pages else "[PDF: no extractable text]"
    except Exception as e:
        return f"[PDF extraction error: {e}]"


def extract_docx_text(data: bytes) -> str:
    """Extract text from a Word .docx file."""
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs) if paragraphs else "[DOCX: no extractable text]"
    except Exception as e:
        return f"[DOCX extraction error: {e}]"


def extract_xlsx_text(data: bytes) -> str:
    """Extract text from an Excel .xlsx file."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                sheets.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(sheets) if sheets else "[XLSX: no data]"
    except Exception as e:
        return f"[XLSX extraction error: {e}]"


def extract_csv_text(data: bytes) -> str:
    """Extract text from a CSV file."""
    try:
        text = data.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = []
        for row in reader:
            if any(row):
                rows.append(" | ".join(row))
        return "\n".join(rows) if rows else "[CSV: no data]"
    except Exception as e:
        return f"[CSV extraction error: {e}]"


def extract_html_text(data: bytes, url: str = "") -> str:
    """Extract readable text from an HTML/ASPX web page."""
    try:
        from bs4 import BeautifulSoup

        html_str = data.decode("utf-8", errors="replace")
        soup = BeautifulSoup(html_str, "html.parser")

        # Remove non-content tags
        for tag in soup(["script", "style", "nav", "footer", "header", "aside", "noscript"]):
            tag.decompose()

        # Try to find the main content area first
        main = (
            soup.find("main")
            or soup.find("article")
            or soup.find("div", {"id": re.compile(r"content|main|body|circular", re.I)})
            or soup.find("div", {"class": re.compile(r"content|main|body|circular", re.I)})
            or soup.body
            or soup
        )

        text = main.get_text(separator="\n", strip=True)

        # Clean up excessive blank lines
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        result = "\n".join(lines)
        return result if result else "[HTML page: no extractable text]"
    except ImportError:
        # Fallback: strip HTML tags with regex if bs4 is not available
        html_str = data.decode("utf-8", errors="replace")
        text = re.sub(r'<[^>]+>', ' ', html_str)
        text = re.sub(r'\s+', ' ', text).strip()
        return text if text else "[HTML page: no extractable text]"
    except Exception as e:
        return f"[HTML extraction error: {e}]"


def extract_attachment_text(filename: str, content_type: str, data: bytes) -> str:
    """Extract text from an attachment based on its type."""
    fname_lower = filename.lower()

    # PDF
    if content_type == "application/pdf" or fname_lower.endswith(".pdf"):
        return extract_pdf_text(data)

    # Word
    if content_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ) or fname_lower.endswith((".docx", ".doc")):
        if fname_lower.endswith(".docx"):
            return extract_docx_text(data)
        return f"[{filename}: .doc format not supported — only .docx is supported]"

    # Excel
    if content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ) or fname_lower.endswith((".xlsx", ".xls")):
        if fname_lower.endswith(".xlsx"):
            return extract_xlsx_text(data)
        return f"[{filename}: .xls format not supported — only .xlsx is supported]"

    # CSV
    if content_type == "text/csv" or fname_lower.endswith(".csv"):
        return extract_csv_text(data)

    # HTML / web page
    if content_type in ("text/html", "application/xhtml+xml") or fname_lower.endswith(
        (".html", ".htm", ".aspx", ".asp", ".php", ".jsp")
    ):
        return extract_html_text(data)

    # Plain text
    if content_type.startswith("text/"):
        return data.decode("utf-8", errors="replace")

    return f"[{filename}: unsupported format ({content_type})]"


# ── Hyperlinked Document Extraction ───────────────────────────────────

# Document extensions we look for in URLs
_DOC_EXTENSIONS = (".pdf", ".docx", ".doc", ".xlsx", ".xls", ".csv", ".txt", ".zip")

# Known exchange / regulatory base domains whose URLs are always worth fetching
_EXCHANGE_BASE_DOMAINS = frozenset({
    "nseindia.com", "bseindia.com", "sebi.gov.in",
    "mcxindia.com", "cdslindia.com", "nsdl.co.in",
})

# Regex to find URLs in plain text and HTML
_URL_PATTERN = re.compile(
    r'https?://[^\s<>"\')]+',
    re.IGNORECASE,
)


def _is_document_url(url: str) -> bool:
    """Check if a URL points to a downloadable document."""
    parsed = urlparse(url.rstrip("/"))
    path = parsed.path.lower()
    return any(path.endswith(ext) for ext in _DOC_EXTENSIONS)


def _is_exchange_url(url: str) -> bool:
    """Check if URL belongs to a known exchange/regulatory domain."""
    host = (urlparse(url).hostname or "").lower()
    return any(host == d or host.endswith(f".{d}") for d in _EXCHANGE_BASE_DOMAINS)


def _filename_from_url(url: str) -> str:
    """Extract a filename from a URL path."""
    parsed = urlparse(url)
    name = parsed.path.rstrip("/").split("/")[-1]
    return unquote(name) if name else "linked_document"


def _download_document(url: str, timeout: int = 30) -> Tuple[bytes, str]:
    """Download a document from a URL. Returns (raw_bytes, content_type)."""
    parsed = urlparse(url)
    domain = (parsed.hostname or "").lower()
    referer = f"{parsed.scheme}://{domain}/"

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        ),
        "Accept": (
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "application/pdf,application/octet-stream,*/*;q=0.8"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Referer": referer,
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Upgrade-Insecure-Requests": "1",
        "sec-ch-ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    session = http_requests.Session()

    # For sites that require cookies/session — visit the domain first
    needs_session = any(
        domain.endswith(d) for d in ("mcxindia.com", "bseindia.com", "cdslindia.com")
    )
    if needs_session:
        try:
            session.get(referer, headers=headers, timeout=10)
        except Exception:
            pass

    resp = session.get(url, headers=headers, timeout=timeout, allow_redirects=True)
    resp.raise_for_status()

    ct = resp.headers.get("Content-Type", "application/octet-stream")
    content_type = ct.split(";")[0].strip().lower()

    return resp.content, content_type


# Public alias — same function, importable from other modules without
# reaching into a "_private" name. Used by consumers/kafka_consumer.py.
def download_document(url: str, timeout: int = 30) -> Tuple[bytes, str]:
    """Download a document from a URL. Returns (raw_bytes, content_type)."""
    return _download_document(url, timeout=timeout)


_CONTENT_TYPE_MAP = {
    "pdf": "application/pdf",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "doc": "application/msword",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "xls": "application/vnd.ms-excel",
    "csv": "text/csv",
    "txt": "text/plain",
    "html": "text/html",
    "htm": "text/html",
    "aspx": "text/html",
    "asp": "text/html",
    "php": "text/html",
}


def extract_linked_documents(text: str) -> List[Dict[str, Any]]:
    """
    Scan text (email body or HTML) for document URLs and exchange web pages,
    download them, and extract text content. Returns list of attachment-style dicts.
    """
    urls = _URL_PATTERN.findall(text)
    target_urls: List[str] = []
    seen = set()

    for url in urls:
        # Clean trailing punctuation that may have been captured
        url = url.rstrip(".,;:!?)")
        if url in seen:
            continue
        # Match document file URLs OR web pages from known exchange domains
        if _is_document_url(url) or _is_exchange_url(url):
            seen.add(url)
            target_urls.append(url)

    if not target_urls:
        return []

    linked_attachments: List[Dict[str, Any]] = []

    for doc_idx, url in enumerate(target_urls):
        filename = _filename_from_url(url)
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
        is_doc_url = _is_document_url(url)
        label = filename if is_doc_url else f"{filename} (web page)"
        print(f"   🔗 Downloading: {label}")
        print(f"      URL: {url}")

        try:
            data, resp_ct = _download_document(url)
            is_html = resp_ct in ("text/html", "application/xhtml+xml")

            # ── ZIP file (only if server didn't return an HTML error page) ──
            if ext == "zip" and not is_html:
                import zipfile
                zip_group = f"zip_{filename}"
                print(f"      📦 Unzipping {filename} in-memory...")
                try:
                    with zipfile.ZipFile(io.BytesIO(data)) as z:
                        for zinfo in z.infolist():
                            if not zinfo.is_dir():
                                zext = zinfo.filename.lower().rsplit(".", 1)[-1] if "." in zinfo.filename else ""
                                if f".{zext}" in _DOC_EXTENSIONS and zext != "zip":
                                    zdata = z.read(zinfo.filename)
                                    zcontent_type = _CONTENT_TYPE_MAP.get(zext, "application/octet-stream")
                                    zextracted = extract_attachment_text(zinfo.filename, zcontent_type, zdata)

                                    linked_attachments.append({
                                        "filename": zinfo.filename,
                                        "type": zcontent_type,
                                        "content": zextracted,
                                        "size": len(zdata),
                                        "source_url": f"{url} [{zinfo.filename}]",
                                        "raw_bytes": zdata,
                                        "doc_group": zip_group,
                                    })
                                    print(f"        ✅ Extracted ZIP item: {zinfo.filename} ({len(zdata)} bytes)")
                except Exception as zip_e:
                    print(f"      ⚠️  Failed to extract zip contents: {zip_e}")
                continue

            # ── HTML / web page response ────────────────────────────────────
            if is_html:
                extracted = extract_html_text(data, url)
                page_name = filename
                if ext not in ("html", "htm", "aspx", "asp", "php", "jsp"):
                    page_name = f"{filename}.html" if filename != "linked_document" else f"page_{doc_idx}.html"
                linked_attachments.append({
                    "filename": page_name,
                    "type": "text/html",
                    "content": extracted,
                    "size": len(data),
                    "source_url": url,
                    "doc_group": f"page_{doc_idx}",
                })
                print(f"      🌐 Extracted web page ({len(data)} bytes)")
                continue

            # ── Document file (PDF, DOCX, XLSX, etc.) ──────────────────────
            # Use response content type if extension is unknown
            content_type = _CONTENT_TYPE_MAP.get(ext, resp_ct)

            # If URL had no extension but server returned a known type, fix filename
            if not ext or ext not in _CONTENT_TYPE_MAP:
                for fext, fct in _CONTENT_TYPE_MAP.items():
                    if fct == resp_ct:
                        filename = f"{filename}.{fext}" if filename != "linked_document" else f"doc_{doc_idx}.{fext}"
                        content_type = fct
                        break

            extracted = extract_attachment_text(filename, content_type, data)
            linked_attachments.append({
                "filename": filename,
                "type": content_type,
                "content": extracted,
                "size": len(data),
                "source_url": url,
                "raw_bytes": data,
                "doc_group": f"link_{doc_idx}",
            })
            print(f"      ✅ Extracted ({len(data)} bytes)")

        except Exception as e:
            linked_attachments.append({
                "filename": filename,
                "type": "unknown",
                "content": f"[Failed to download {filename}: {e}]",
                "size": 0,
                "source_url": url,
                "doc_group": f"link_{doc_idx}",
            })
            print(f"      ⚠️  Download failed: {e}")

    return linked_attachments
